"""Release checks for the source-attributed public capability samples."""

from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
ALLOWED_SOURCE_HOSTS = {
    "en.dmgmori.com",
    "us.dmgmori.com",
    "www.gleason.com",
    "www.iscar.com",
    "www.kapp-niles.com",
}


def _assert_safe_workbook(path: Path, data_sheet: str, source_header: str) -> None:
    workbook = load_workbook(path, data_only=False, read_only=True)
    assert data_sheet in workbook.sheetnames
    assert "README" in workbook.sheetnames
    sheet = workbook[data_sheet]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    source_col = headers.index(source_header) + 1
    checked_col = headers.index("Source checked on") + 1
    rows = list(sheet.iter_rows(min_row=2))
    assert rows
    for row in rows:
        source = row[source_col - 1].value
        checked = row[checked_col - 1].value
        assert isinstance(source, str) and urlparse(source).scheme == "https"
        assert urlparse(source).hostname in ALLOWED_SOURCE_HOSTS
        assert checked is not None
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            assert all(cell.data_type != "f" for cell in row), (
                f"Formula found in public data: {worksheet.title}"
            )


def test_public_machine_data_is_attributed_and_formula_free() -> None:
    _assert_safe_workbook(ROOT / "data" / "machines.xlsx", "Export", "Capability source URL")


def test_public_tool_data_is_attributed_and_formula_free() -> None:
    _assert_safe_workbook(ROOT / "data" / "tools.xlsx", "Tool_Selection", "Source URL")
