"""Process card export regression tests: every operation should show its matched machine/tool/status."""

import uuid
from unittest import mock

import pandas as pd
from openpyxl import load_workbook

from service import PlanningService
from workflow import JobStore


def _service_without_rag() -> PlanningService:
    """Build a service that does not write to the real RAG case library (export tests must not touch rag/data)."""
    service = object.__new__(PlanningService)
    service.store = JobStore()
    service._store_exported_card_to_rag = mock.Mock()
    return service


def _make_result() -> dict:
    """Build a result close to real workflow output (including matched machines/tools)."""
    machine = {
        "designation": "QT Nexus 200 Universal 500", "manufacturer": "Yamazaki Mazak",
        "unique_identifier": "CP-TURN-001", "turning_length_mm": 511,
        "max_turning_diameter_rod_mm": 65, "max_turning_diameter_chuck_mm": None,
    }
    tool = {
        "cutting_tool_grade": "IC808", "machining_process": "ISO Turning",
        "first_choice": True, "hard_tough_rank": 1,
    }
    operations = [
        {"operation_no": 1, "name": "Blanking", "stage": "blank",
         "description": "Cut from bar stock.", "process_category": None,
         "feature_id": None, "conditional": False},
        {"operation_no": 2, "name": "Face Turning", "stage": "datum",
         "description": "Turn both faces.", "process_category": "ISO Turning",
         "feature_id": None, "conditional": False},
        {"operation_no": 3, "name": "Rough Turning", "stage": "rough",
         "description": "Rough turn profile.", "process_category": "ISO Turning",
         "feature_id": None, "conditional": False},
    ]
    operation_resources = [
        {"operation_no": 1, "operation_name": "Blanking", "process_category": None,
         "verification_status": "not_applicable",
         "machine_recommendations": [], "tool_recommendations": [],
         "note": "No tool/equipment verification needed."},
        {"operation_no": 2, "operation_name": "Face Turning", "process_category": "ISO Turning",
         "verification_status": "satisfied",
         "machine_recommendations": [machine], "tool_recommendations": [tool],
         "note": "Machine: Found turning machine records matching size."},
        {"operation_no": 3, "operation_name": "Rough Turning", "process_category": "ISO Turning",
         "verification_status": "satisfied",
         "machine_recommendations": [machine], "tool_recommendations": [tool],
         "note": "Machine: Found turning machine records matching size."},
    ]
    return {
        "process_route": operations,
        "geometry": {
            "total_length_mm": 180.0, "blank_diameter_mm": 65.0,
            "max_finished_diameter_mm": 50.0,
            "segments": [{"segment_id": "S01", "diameter_mm": 50, "length_mm": 180.0,
                          "global_start_mm": 0, "global_end_mm": 180.0}],
            "features": [],
        },
        "resource_selection": {
            "operation_resources": operation_resources,
            "turning_machine_candidates": [machine],
            "scope_note": "Current tool table verifies cutting-tool grades only.",
        },
        "heat_treatment_decision": {},
    }


def _read_card(path) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name="Process Card", header=None)


class TestProcessCardEquipmentColumns:
    def test_machine_tool_status_columns_show_matched_resources(self, tmp_path):
        """Regression: Machine/Tool/Status columns should read machine_recommendations /
        tool_recommendations / verification_status (previously read non-existent machine/tool/status keys, producing all "-")."""
        service = _service_without_rag()
        job_id = uuid.uuid4().hex[:12]
        service.store.create(job_id, {
            "material": "45", "blank_diameter_mm": 65, "blank_type": "solid",
            "global_requirements": {"heat_treatment": "none", "surface_treatment": "none",
                                    "batch_quantity": 1},
        })
        service.store.update(job_id, status="completed", result=_make_result())

        path = service.export_process_card_excel(job_id)
        try:
            df = _read_card(path)
            # Find the Process Route & Equipment header row
            header_row = df[df[0] == "Op#"].index[0]
            rows = df.iloc[header_row + 1:].copy()
            rows = rows[rows[0].notna()].astype(str)

            face_row = rows[rows[1].str.contains("Face Turning")].iloc[0]
            assert "QT Nexus 200 Universal 500" in face_row[4], (
                f"Machine cell should show matched designation, got {face_row[4]!r}"
            )
            assert "IC808" in face_row[5], f"Tool cell should show cutting grade, got {face_row[5]!r}"
            assert face_row[6] == "satisfied", f"Status cell should show verification_status, got {face_row[6]!r}"

            blank_row = rows[rows[1].str.contains("Blanking")].iloc[0]
            assert blank_row[4] == "-", "not_applicable operations should not have equipment"
        finally:
            path.unlink(missing_ok=True)

    def test_no_standalone_turning_machine_candidates_section(self, tmp_path):
        """Regression: no standalone Turning Machine Candidates section (equipment is now shown under each operation)."""
        service = _service_without_rag()
        job_id = uuid.uuid4().hex[:12]
        service.store.create(job_id, {"global_requirements": {}})
        service.store.update(job_id, status="completed", result=_make_result())

        path = service.export_process_card_excel(job_id)
        try:
            df = _read_card(path)
            assert not (df == "Turning Machine Candidates").any().any(), (
                "Process card should not contain a standalone Turning Machine Candidates section"
            )
        finally:
            path.unlink(missing_ok=True)


def _route_rows(path):
    """Extract rows of the Process Route & Equipment section from the process card."""
    df = _read_card(path)
    header_row = df[df[0] == "Op#"].index[0]
    rows = df.iloc[header_row + 1:].copy()
    return rows[rows[0].notna()].astype(str)


class TestCustomRouteExport:
    def test_formula_like_operation_text_is_exported_as_literal_text(self):
        service = _service_without_rag()
        job_id = uuid.uuid4().hex[:12]
        service.store.create(job_id, {"global_requirements": {}})
        result = _make_result()
        result["process_route"][0]["name"] = '=HYPERLINK("https://example.invalid","click")'
        service.store.update(job_id, status="completed", result=result)

        path = service.export_process_card_excel(job_id)
        try:
            workbook = load_workbook(path, data_only=False)
            cells = [cell for row in workbook["Process Card"].iter_rows() for cell in row]
            matching = [cell for cell in cells if isinstance(cell.value, str) and "HYPERLINK" in cell.value]
            assert len(matching) == 1
            assert matching[0].data_type == "s"
            assert matching[0].value.startswith("'=")
        finally:
            path.unlink(missing_ok=True)

    def test_custom_route_reorders_operations_and_keeps_resources(self, tmp_path):
        """Custom route exports in user order with sequential Op# numbering; machine/tool/status still
        associate by original operation_no (the stable resource key)."""
        service = _service_without_rag()
        job_id = uuid.uuid4().hex[:12]
        service.store.create(job_id, {
            "material": "45", "blank_diameter_mm": 65, "blank_type": "solid",
            "global_requirements": {"heat_treatment": "none", "surface_treatment": "none",
                                    "batch_quantity": 1},
        })
        ops = _make_result()["process_route"]
        # Reorder to [Rough Turning(3), Blanking(1), Face Turning(2)]; operation_no stays as the stable key
        custom_route = [dict(ops[2]), dict(ops[0]), dict(ops[1])]
        service.store.update(job_id, status="completed", result=_make_result(),
                             custom_route=custom_route)

        path = service.export_process_card_excel(job_id)
        try:
            rows = _route_rows(path)
            assert list(rows[1]) == ["Rough Turning", "Blanking", "Face Turning"], (
                "Operations should be output in custom order"
            )
            assert list(rows[0]) == ["1", "2", "3"], "Op# should be numbered consecutively in current order"

            rough_row = rows[rows[1].str.contains("Rough Turning")].iloc[0]
            assert "QT Nexus 200 Universal 500" in rough_row[4], (
                f"Rough Turning should keep its machine from original operation_no=3 after reordering, got {rough_row[4]!r}"
            )
            assert "IC808" in rough_row[5]
            assert rough_row[6] == "satisfied"

            blank_row = rows[rows[1].str.contains("Blanking")].iloc[0]
            assert blank_row[4] == "-", "Blanking (not_applicable) should not have equipment"
        finally:
            path.unlink(missing_ok=True)

    def test_reset_custom_route_falls_back_to_original(self, tmp_path):
        """After custom_route is cleared (reset), the process card falls back to the workflow-generated original order."""
        service = _service_without_rag()
        job_id = uuid.uuid4().hex[:12]
        service.store.create(job_id, {"global_requirements": {}})
        service.store.update(job_id, status="completed", result=_make_result(),
                             custom_route=None)

        path = service.export_process_card_excel(job_id)
        try:
            rows = _route_rows(path)
            assert list(rows[1]) == ["Blanking", "Face Turning", "Rough Turning"]
        finally:
            path.unlink(missing_ok=True)


class TestCustomizeRouteService:
    def test_customize_and_reset_route_methods(self):
        """customize_route saves the adjusted route; reset_custom_route clears it back to None."""
        from models.process import ProcessOperation

        service = _service_without_rag()
        job_id = uuid.uuid4().hex[:12]
        service.store.create(job_id, {"global_requirements": {}})
        ops = _make_result()["process_route"]
        service.store.update(job_id, status="completed", result=_make_result())

        reordered = [ProcessOperation(**ops[2]), ProcessOperation(**ops[0]), ProcessOperation(**ops[1])]
        saved = service.customize_route(job_id, reordered)
        assert [o["operation_no"] for o in saved] == [3, 1, 2]
        assert service.store.get(job_id)["custom_route"][0]["name"] == "Rough Turning"

        service.reset_custom_route(job_id)
        assert service.store.get(job_id)["custom_route"] is None

    def test_customize_route_requires_generated_route(self):
        """Customization is not allowed when result is not ready (no process_route)."""
        import pytest

        service = _service_without_rag()
        job_id = uuid.uuid4().hex[:12]
        service.store.create(job_id, {"global_requirements": {}})
        service.store.update(job_id, status="failed", result={"traceback": "boom"})

        from models.process import ProcessOperation
        ops = _make_result()["process_route"]
        with pytest.raises(ValueError):
            service.customize_route(job_id, [ProcessOperation(**ops[0])])
