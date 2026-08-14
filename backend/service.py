"""Planning service layer: job creation, HITL resume, thread pool management."""

from __future__ import annotations

import hashlib
import json
import logging
import os
import signal
import threading
import time
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
from typing import Any

from langgraph.types import Command

from models.workflow import PlanningRequest
from models.input import ChoicesRequest
from workflow import JobStore, Workflow

logger = logging.getLogger(__name__)

HEARTBEAT_TIMEOUT = int(os.getenv("HEARTBEAT_TIMEOUT", "30"))

# ---- A1 job-level cache (demo/development) ----
# Identical requests (same hash) reuse the previous workflow result, skipping the whole
# LangGraph run (about 56s -> sub-second).
# In-memory, in-process, cleared on restart; during demos re-submitting the same shaft
# returns results immediately.
# Disable/tune via .env: JOB_CACHE_ENABLED=false / JOB_CACHE_TTL_SECONDS / JOB_CACHE_MAX_ENTRIES.
JOB_CACHE_ENABLED = (
    os.getenv("JOB_CACHE_ENABLED", "true").strip().lower()
    in ("1", "true", "yes", "on")
)
JOB_CACHE_TTL_SECONDS = int(os.getenv("JOB_CACHE_TTL_SECONDS", "3600"))
JOB_CACHE_MAX_ENTRIES = int(os.getenv("JOB_CACHE_MAX_ENTRIES", "50"))

# Concurrency lock for exporting the process card -> syncing to the RAG case library
_RAG_CASE_LOCK = threading.Lock()


class JobCache:
    """Thread-safe in-memory job result cache (TTL + max-entry eviction, demo/development only).

    The key is a deterministic hash of the request payload; the value is a deep copy of the
    final workflow result, so multiple jobs never share the same object and pollute each other.
    """

    def __init__(self, max_entries: int = JOB_CACHE_MAX_ENTRIES,
                 ttl_seconds: float = JOB_CACHE_TTL_SECONDS) -> None:
        self._max_entries = max_entries
        self._ttl = ttl_seconds
        self._data: dict[str, dict[str, Any]] = {}
        self._lock = threading.Lock()
        self.enabled = max_entries > 0 and ttl_seconds > 0

    def get(self, key: str) -> dict[str, Any] | None:
        if not self.enabled:
            return None
        with self._lock:
            entry = self._data.get(key)
            if entry is None:
                return None
            if time.time() - entry["stored_at"] > self._ttl:
                del self._data[key]
                return None
            return deepcopy(entry["result"])

    def put(self, key: str, result: dict[str, Any]) -> None:
        if not self.enabled:
            return
        with self._lock:
            # Evict the oldest entry (by write time) when over the limit
            if key not in self._data and len(self._data) >= self._max_entries:
                oldest_key = min(self._data, key=lambda k: self._data[k]["stored_at"])
                del self._data[oldest_key]
            self._data[key] = {"result": deepcopy(result), "stored_at": time.time()}

    def clear(self) -> None:
        with self._lock:
            self._data.clear()

    def __len__(self) -> int:
        return len(self._data)


def _request_cache_key(payload: dict[str, Any]) -> str:
    """Deterministic hash of the request payload: canonical JSON with sorted keys -> sha256.

    The same input always yields the same key regardless of dict key order, implementing
    "same request -> reused result".
    """
    canonical = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


class PlanningService:
    """Process planning service: manages the job lifecycle."""

    def __init__(self) -> None:
        self.store = JobStore()
        self.workflow = Workflow(self.store)
        self.job_cache = JobCache()
        max_workers = min(max(2, os.cpu_count() or 4), 8)
        self.executor = ThreadPoolExecutor(max_workers=max_workers)

        # Heartbeat watchdog: auto-shutdown if no browser heartbeat within timeout
        self.last_heartbeat = time.time()
        self._watchdog_active = True
        self._watchdog_thread = threading.Thread(target=self._watchdog, daemon=True)
        self._watchdog_thread.start()

    def touch_heartbeat(self) -> None:
        self.last_heartbeat = time.time()

    def _watchdog(self) -> None:
        while self._watchdog_active:
            time.sleep(5)
            elapsed = time.time() - self.last_heartbeat
            if elapsed > HEARTBEAT_TIMEOUT:
                logger.warning("No heartbeat for %.0fs, shutting down...", elapsed)
                os.kill(os.getpid(), signal.SIGTERM)
                self._watchdog_active = False
                return

    def create(self, request: PlanningRequest) -> str:
        job_id = uuid.uuid4().hex[:12]
        payload = request.model_dump(mode="json")
        cache_key = _request_cache_key(payload)

        # A1 cache hit: skip the whole workflow and reuse the previous result (demo: same shaft returns in seconds)
        if self.job_cache.enabled:
            cached = self.job_cache.get(cache_key)
            if cached is not None:
                logger.info("[%s] Job cache HIT (%s), reusing previous result.",
                            job_id, cache_key[:8])
                self.store.create(job_id, payload)
                self.store.update(
                    job_id,
                    status="completed", progress=100, current_step="completed",
                    pending_choices=[], message="Cache hit: input identical to the previous run, reusing the previous process result. "
                                                 "(To force recomputation, set JOB_CACHE_ENABLED=false in .env)",
                    result=cached, _cache_key=cache_key,
                )
                return job_id

        self.store.create(job_id, payload)
        try:
            self.executor.submit(self._initial, job_id, payload, cache_key)
        except Exception as error:
            self.store.update(job_id, status="failed", progress=100, current_step="failed", message=f"Task submission failed: {error}")
        return job_id

    def _initial(self, job_id: str, payload: dict[str, Any], cache_key: str | None = None) -> None:
        self._invoke(job_id, {"job_id": job_id, "request": payload, "user_choices": {}, "route_hashes": [], "repair_count": 0, "status": "running"}, cache_key=cache_key)

    def resume(self, job_id: str, choices: ChoicesRequest) -> None:
        current = self.store.get(job_id)
        if current["status"] != "waiting_user_choice":
            raise ValueError("Task is not in waiting for user choice state.")
        self.store.update(job_id, status="running", message="Choices received, continuing.", pending_choices=[])
        self.executor.submit(self._invoke, job_id, Command(resume=choices.model_dump(mode="json")))

    def customize_route(self, job_id: str, operations: list[Any]) -> list[dict[str, Any]]:
        """Save the user-customized process route (adjustment stage before the process card is generated).

        Each operation keeps its original operation_no as a stable resource key; when the process card
        is exported, machine/tool recommendations are linked by that key. Returns the saved route JSON.
        """
        job = self.store.get(job_id)
        if job["result"] is None or not job["result"].get("process_route"):
            raise ValueError("Process route is not ready; cannot customize.")
        numbers = [op.operation_no for op in operations]
        if len(numbers) != len(set(numbers)):
            raise ValueError("operation_no must be unique.")
        payload = [op.model_dump(mode="json") for op in operations]
        self.store.update(job_id, custom_route=payload)
        return payload

    def reset_custom_route(self, job_id: str) -> None:
        """Clear the customized route and restore the original route generated by the workflow."""
        job = self.store.get(job_id)
        if job["result"] is None:
            raise ValueError("Task result not ready.")
        # JobStore only has a merge-style update; None means "not customized", and all readers fall back to the original route
        self.store.update(job_id, custom_route=None)

    def _invoke(self, job_id: str, graph_input: dict[str, Any] | Command,
                cache_key: str | None = None) -> None:
        from langgraph.errors import GraphInterrupt
        config = {"configurable": {"thread_id": job_id}}
        try:
            result = self.workflow.graph.invoke(graph_input, config=config)
            if result.get("__interrupt__"):
                logger.info("[%s] HITL interrupt captured, waiting for user choice.", job_id)
                return
            final_status = result.get("status", "completed")
            if final_status == "resource_mismatch":
                message, current_step = "Critical turning resources not satisfied, process terminated.", "resource_mismatch"
            elif final_status == "failed":
                message, current_step = result.get("verification", {}).get("message") or "Task execution failed.", "failed"
            else:
                message, current_step = result.get("verification", {}).get("message") or "Task completed.", "completed"
            update_values: dict[str, Any] = {"result": result, "status": final_status, "message": message, "current_step": current_step, "pending_choices": []}
            if final_status in {"resource_mismatch", "failed", "completed"}:
                update_values["progress"] = 100
                # A1 cache write-back: only written on the clean initial request path; HITL resume/failure is not written.
                # Later identical requests reuse the result directly, skipping the whole workflow.
                if cache_key and self.job_cache.enabled \
                        and final_status in {"completed", "resource_mismatch"}:
                    self.job_cache.put(cache_key, result)
                    logger.info("[%s] Job cache store (%s) — %d entries.",
                                job_id, cache_key[:8], len(self.job_cache))
            self.store.update(job_id, **update_values)
        except GraphInterrupt:
            logger.info("[%s] HITL interrupt: waiting for user precision choice.", job_id)
            return
        except Exception as error:
            self.store.update(job_id, status="failed", progress=100, current_step="failed",
                              message="Backend execution failed.", error=f"{type(error).__name__}: {error}",
                              result={"traceback": traceback.format_exc()})

    def export_process_card_excel(self, job_id: str) -> Path:
        """Generate the process card Excel file, save it to the project output/ directory, and return the file path."""
        from pathlib import Path as _Path
        from openpyxl import Workbook
        from rules import HEAT_NAME, SURFACE_NAME, FEATURE_NAME
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        job = self.store.get(job_id)
        if job["status"] not in {"completed", "resource_mismatch", "failed"}:
            raise ValueError("Task not yet completed.")
        if job["result"] is None:
            raise ValueError("Task result not ready.")

        result = job["result"]
        request_data = job.get("request") or {}
        # The user-customized route takes priority; operation_no is the stable resource key, so resources still follow the operation after reordering
        route = job.get("custom_route") or result.get("process_route", [])
        geometry = result.get("geometry", {})
        resources = result.get("resource_selection", {})
        req_global = request_data.get("global_requirements", {})
        heat_decision = result.get("heat_treatment_decision", {})
        heat_label = heat_decision.get("process_name") or HEAT_NAME.get(req_global.get("heat_treatment", "none"), "None")
        heat_requirements = []
        if req_global.get("target_hardness_hrc") is not None:
            heat_requirements.append(f"{req_global['target_hardness_hrc']} HRC")
        if req_global.get("case_depth_mm") is not None:
            heat_requirements.append(f"case depth {req_global['case_depth_mm']} mm")

        op_resources_map = {}
        for rr in resources.get("operation_resources", []):
            op_resources_map[rr.get("operation_no", 0)] = rr

        wb = Workbook()
        ws = wb.active
        ws.title = "Process Card"

        # ── Styles ──
        hdr_font = Font(bold=True, size=11, color="FFFFFF")
        hdr_fill = PatternFill(start_color="4F6EF7", end_color="4F6EF7", fill_type="solid")
        sub_font = Font(bold=True, size=10)
        sub_fill = PatternFill(start_color="E8EDFF", end_color="E8EDFF", fill_type="solid")
        title_font = Font(bold=True, size=14, color="1E293B")
        section_font = Font(bold=True, size=12, color="4F6EF7")
        bdr = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lw = Alignment(vertical="center", wrap_text=True)

        def hdr_row(row, ncol):
            for c in range(1, ncol + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = bdr

        def body_rng(r1, r2, ncol, left_col=0):
            for rr in range(r1, r2 + 1):
                for c in range(1, ncol + 1):
                    cell = ws.cell(row=rr, column=c)
                    cell.border = bdr
                    cell.alignment = lw if (left_col and c == left_col) else ctr

        def wr(row, vals):
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)

        r = 1
        # ═══════════════════════════════════════
        # Title
        # ═══════════════════════════════════════
        ws.merge_cells("A1:H1")
        ws["A1"] = "ShaftPlanner - Process Card"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center")

        # ═══════════════════════════════════════
        # Part Information
        # ═══════════════════════════════════════
        r = 3
        ws.merge_cells(f"A{r}:H{r}")
        ws.cell(row=r, column=1, value="Part Information").font = section_font
        r += 1
        info = [
            ("Job ID", job_id),
            ("Material", request_data.get("material", "-")),
            ("Blank Type", request_data.get("blank_type", "solid")),
            ("Blank Diameter", f"φ{request_data.get('blank_diameter_mm', geometry.get('blank_diameter_mm', '-'))} mm"),
            ("Total Length", f"{geometry.get('total_length_mm', '-')} mm"),
            ("Max Finished Dia", f"φ{geometry.get('max_finished_diameter_mm', '-')} mm"),
            ("Heat Treatment", heat_label),
            ("Heat-treatment Requirements", ", ".join(heat_requirements) if heat_requirements else "Drawing/specification confirmation required" if heat_label != "None" else "-"),
            ("Surface Treatment", SURFACE_NAME.get(req_global.get("surface_treatment", "none"), "None")),
            ("Batch Quantity", req_global.get("batch_quantity", 1)),
        ]
        for param, val in info:
            wr(r, [param, str(val)])
            ws.cell(row=r, column=1).font = sub_font
            ws.cell(row=r, column=1).fill = sub_fill
            ws.cell(row=r, column=2).alignment = lw
            r += 1
        body_rng(4, r - 1, 2)

        # ═══════════════════════════════════════
        # Segments
        # ═══════════════════════════════════════
        segments = geometry.get("segments", [])
        r += 1
        ws.merge_cells(f"A{r}:H{r}")
        ws.cell(row=r, column=1, value="Stepped Shaft Segments").font = section_font
        r += 1
        wr(r, ["Seg ID", "Diameter (mm)", "Length (mm)", "Start Pos (mm)", "End Pos (mm)",
               "Upper Dev (mm)", "Lower Dev (mm)", "Ra (μm)"])
        hdr_row(r, 8)
        r += 1
        seg_start = r
        for seg in segments:
            wr(r, [
                seg.get("segment_id", ""), seg.get("diameter_mm", ""), seg.get("length_mm", ""),
                seg.get("global_start_mm", ""), seg.get("global_end_mm", ""),
                seg.get("diameter_upper_deviation_mm", ""), seg.get("diameter_lower_deviation_mm", ""),
                seg.get("roughness_ra", ""),
            ])
            r += 1
        body_rng(seg_start, r - 1, 8)

        # ═══════════════════════════════════════
        # Features
        # ═══════════════════════════════════════
        features = geometry.get("features", [])
        r += 1
        ws.merge_cells(f"A{r}:H{r}")
        ws.cell(row=r, column=1, value="Conditional Features").font = section_font
        r += 1
        wr(r, ["Feature ID", "Type", "Segment", "Position (mm)", "Tolerance (mm)", "Ra (μm)", "Precision", "Parameters"])
        hdr_row(r, 8)
        r += 1
        feat_start = r
        for feat in features:
            ft = feat.get("feature_type", "")
            tol_str = f"{feat.get('tolerance_lower_mm', '-')} / {feat.get('tolerance_upper_mm', '-')}"
            params = []
            if ft == "keyway":
                params.append(f"W:{feat.get('keyway_width_mm','-')} D:{feat.get('keyway_depth_mm','-')} L:{feat.get('feature_length_mm','-')}")
            elif ft == "hole":
                params.append(f"φ{feat.get('hole_diameter_mm','-')} {feat.get('hole_type','through')} {feat.get('hole_direction','radial')}")
            elif ft == "thread":
                params.append(f"{feat.get('thread_specification','-')} {feat.get('thread_handedness','right')}")
            elif ft == "bearing_seat":
                params.append(f"φ{feat.get('bearing_seat_diameter_mm','-')} {feat.get('bearing_seat_tolerance','IT6')}")
            elif ft == "spline":
                params.append(f"{feat.get('spline_type','-')} Z:{feat.get('spline_teeth','-')} m:{feat.get('spline_module','-')}")
            elif ft == "taper":
                params.append(f"1:{feat.get('taper_ratio','-')} D:{feat.get('taper_large_diameter_mm','-')} L:{feat.get('taper_length_mm','-')}")
            elif ft == "groove":
                params.append(f"{feat.get('groove_type','-')} W:{feat.get('groove_width_mm','-')} D:{feat.get('groove_depth_mm','-')}")
            elif ft == "seal_area":
                params.append(f"{feat.get('seal_type','-')} φ{feat.get('seal_diameter_mm','-')}")
            elif ft == "gear_teeth":
                params.append(f"m:{feat.get('gear_module','-')} Z:{feat.get('gear_teeth','-')} α:{feat.get('gear_pressure_angle','-')}°")
            elif ft == "flange":
                params.append(f"φ{feat.get('flange_diameter_mm','-')} T:{feat.get('flange_thickness_mm','-')} Holes:{feat.get('flange_holes',0)}")
            wr(r, [
                feat.get("feature_id", ""), FEATURE_NAME.get(ft, ft),
                feat.get("resolved_segment_id", ""), feat.get("global_position_mm", ""),
                tol_str, feat.get("roughness_ra", ""),
                "High" if feat.get("high_precision") else "Normal",
                "; ".join(params) if params else "-",
            ])
            r += 1
        body_rng(feat_start, r - 1, 8, left_col=8)

        # ═══════════════════════════════════════
        # Process Route & Equipment
        # ═══════════════════════════════════════
        r += 1
        ws.merge_cells(f"A{r}:H{r}")
        ws.cell(row=r, column=1, value="Process Route & Equipment").font = section_font
        r += 1
        wr(r, ["Op#", "Operation", "Stage", "Description", "Machine", "Tool", "Status", "Note"])
        hdr_row(r, 8)
        r += 1
        route_start = r
        for idx, op in enumerate(route):
            # Display sequence follows the current route order; resources are linked by the original operation_no (stable key)
            op_no = idx + 1
            res = op_resources_map.get(op.get("operation_no", op_no), {})

            # Machine/tool/status all come from resource_selection.operation_resources.
            # Each operation shows its top recommended machines and tools (instead of a separate candidate list).
            machine_recos = res.get("machine_recommendations", []) or []
            if machine_recos:
                m = machine_recos[0]
                spec_parts = []
                if m.get("turning_length_mm"):
                    spec_parts.append(f"L:{m['turning_length_mm']}mm")
                if m.get("max_turning_diameter_rod_mm"):
                    spec_parts.append(f"Rod:φ{m['max_turning_diameter_rod_mm']}mm")
                if m.get("max_turning_diameter_chuck_mm"):
                    spec_parts.append(f"Chuck:φ{m['max_turning_diameter_chuck_mm']}mm")
                machine_cell = m.get("designation", "") + (
                    " (" + " ".join(spec_parts) + ")" if spec_parts else ""
                )
            else:
                machine_cell = "-"

            tool_recos = res.get("tool_recommendations", []) or []
            tool_cell = ", ".join(
                t.get("cutting_tool_grade", "") for t in tool_recos if t.get("cutting_tool_grade")
            ) or "-"

            status_cell = res.get("verification_status", "") or ""

            wr(r, [
                op_no, op.get("name", ""), op.get("stage", ""), op.get("description", ""),
                machine_cell, tool_cell, status_cell, res.get("note", ""),
            ])
            r += 1
        body_rng(route_start, r - 1, 8, left_col=4)

        # ── Column widths ──
        for c, w in enumerate([14, 22, 18, 42, 28, 28, 16, 32], 1):
            ws.column_dimensions[get_column_letter(c)].width = w

        # ── Save ──
        project_root = _Path(__file__).resolve().parent.parent
        output_dir = project_root / "output"
        output_dir.mkdir(exist_ok=True)
        file_path = output_dir / f"process_card_{job_id}.xlsx"
        wb.save(file_path)
        logger.info("Process card exported: %s", file_path)

        # ---- RAG: store the exported process card into the case library and vectorize it (failure does not affect the export flow) ----
        try:
            self._store_exported_card_to_rag(
                job_id, result, request_data, geometry, route,
                heat_label, op_resources_map,
            )
        except Exception as exc:  # noqa: BLE001 - a RAG storage failure must not block the export
            logger.warning("Failed to store process card into RAG case library: %s", exc)

        return file_path

    def _store_exported_card_to_rag(
        self,
        job_id: str,
        result: dict[str, Any],
        request_data: dict[str, Any],
        geometry: dict[str, Any],
        route: list[dict[str, Any]],
        heat_label: str,
        op_resources_map: dict[int, dict[str, Any]],
    ) -> None:
        """While exporting the process card, store the process data into the RAG case library and vectorize it.

        Steps:
          1. Build the current process card into standard case JSON (fields compatible with rag.splitters.case_splitter)
          2. Write to rag/data/cases/exported_cases.json (deduplicated by case_id)
          3. Rebuild the case library index (chunking -> embedding vectorization -> ChromaDB -> BM25 sync)
        """
        import json as _json
        from pathlib import Path as _Path

        from rag.indexer import IndexBuilder

        with _RAG_CASE_LOCK:
            # ---- 1. Build case data ----
            process_plan = []
            for idx, op in enumerate(route):
                res = op_resources_map.get(op.get("operation_no", idx + 1), {})
                # Consistent with the process card export: the actual fields are machine_recommendations / tool_recommendations
                machine_recos = res.get("machine_recommendations", []) or []
                tool_recos = res.get("tool_recommendations", []) or []
                process_plan.append({
                    "step_no": idx + 1,
                    "name": op.get("name", ""),
                    "stage": op.get("stage", ""),
                    "description": op.get("description", ""),
                    "machine": ", ".join(
                        m.get("designation", "") for m in machine_recos if m.get("designation")
                    ),
                    "tool": ", ".join(
                        t.get("cutting_tool_grade", "") for t in tool_recos if t.get("cutting_tool_grade")
                    ),
                })

            req_global = request_data.get("global_requirements", {})
            case = {
                "case_id": f"EXPORT-{job_id}",
                "part_name": request_data.get("part_name")
                or request_data.get("material")
                or f"Job-{job_id}",
                "material": request_data.get("material", "Unknown"),
                "taxonomy_id": "",
                "industry": "",
                "application": "",
                "heat_treatment": heat_label,
                "tolerance": (
                    f"{req_global['target_hardness_hrc']} HRC"
                    if req_global.get("target_hardness_hrc") is not None
                    else ""
                ),
                "surface_roughness": "",
                "length_mm": geometry.get("total_length_mm", ""),
                "diameter_mm": geometry.get("max_finished_diameter_mm", ""),
                "main_features": geometry.get("features", []),
                "description": request_data.get("notes", ""),
                "process_plan": process_plan,
            }

            # ---- 2. Write the case source file (deduplicated by case_id) ----
            project_root = _Path(__file__).resolve().parent.parent
            cases_file = (
                project_root / "backend" / "rag" / "data" / "cases" / "exported_cases.json"
            )
            cases_file.parent.mkdir(parents=True, exist_ok=True)

            data: dict[str, Any] = {"cases": []}
            if cases_file.exists():
                try:
                    _loaded = _json.loads(cases_file.read_text(encoding="utf-8"))
                    if isinstance(_loaded, dict) and isinstance(_loaded.get("cases"), list):
                        data = _loaded
                except Exception:  # noqa: BLE001 - restart from scratch if the file is corrupted
                    data = {"cases": []}

            case_id = case["case_id"]
            data["cases"] = [c for c in data["cases"] if c.get("case_id") != case_id]
            data["cases"].append(case)
            cases_file.write_text(
                _json.dumps(data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            logger.info("Process card stored to RAG case file: %s", cases_file)

            # ---- 3. Rebuild the case library index (chunking -> vectorization -> ChromaDB + BM25) ----
            count = IndexBuilder().build_case_index()
            logger.info(
                "RAG case index rebuilt: %d chunks after exporting process card %s",
                count, job_id,
            )

    def result(self, job_id: str) -> dict[str, Any]:
        job = self.store.get(job_id)
        if job["status"] not in {"completed", "resource_mismatch", "failed"}:
            raise ValueError("Task not yet completed.")
        if job["result"] is None:
            raise ValueError("Task result is being written, please retry later.")
        result = job["result"]
        return {
            "job_id": job_id, "status": job["status"], "message": job["message"], "error": job["error"],
            "plan": result.get("plan"), "geometry": result.get("geometry"), "capability": result.get("capability"),
            "process_route": result.get("process_route", []), "custom_route": job.get("custom_route"),
            "resource_selection": result.get("resource_selection"),
            "verification": result.get("verification"), "execution_trace": result.get("execution_trace", []), "result": result,
        }
