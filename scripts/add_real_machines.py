#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Write real production machines into data/machines.xlsx (covering all process categories of the engine).

Added (real models + capabilities + source URLs):
- Drilling          : TBT deep hole drills (ML-250 / B3S / BW300)
- Indexable Milling : Mazak VCN vertical machining centers / DMG MORI DMU 50
- Boring            : TOS WHN horizontal boring mills (WHN 13 / WHN 110)
- Thread Grinding   : Matrix thread grinders (37 / 47 / THW-3080)
- Balancing         : Schenck HM horizontal balancing machines (HM2 / HM4 / HM40)
- Shot Blasting     : Wheelabrator shot blast machines (WS-4/210 / CT-2-45/7-430)
Also adds a "Threading|Taper Turning|Grooving" capability to existing lathe rows (lathes can cut threads, tapers and grooves).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "machines.xlsx"
SHEET = "Export"

# New rows: (Designation, Manufacturer, Machine type, Supported processes,
#            Max workpiece length mm, Max workpiece diameter mm, Max workpiece weight kg,
#            Max gear module mm, Capability source URL, Capability notes)
NEW_ROWS = [
    # ── Drilling (deep hole drills) ──
    (
        "TBT ML-250",
        "TBT Tiefbohr-Technik",
        "CNC deep hole drilling machine",
        "Drilling",
        1550,
        100,
        1000,
        None,
        "https://www.surplex.com/en/l/tbt-ml-300-4-1450-deep-hole-drilling-machine-A7-40658-11",
        "Gundrilling (ELB) + BTA; hole diameter up to 100 mm, drilling depth up to 1550 mm, up to 4 spindles.",
    ),
    (
        "TBT B3S",
        "TBT Tiefbohr-Technik",
        "Horizontal BTA deep hole drilling machine",
        "Drilling",
        15000,
        200,
        20000,
        None,
        "https://www.europages.co.uk/en/company/tbt-tiefbohrtechnik-gmbh-co-297880/products/b-serie-horizontale-grossmaschinen-fuer-tiefbohrtechnik-37373436",
        "BTA/STS deep hole drilling, full drilling diameter 20-200 mm, drilling depth 2000-15000 mm.",
    ),
    (
        "TBT BW300",
        "TBT Tiefbohr-Technik",
        "Table deep hole drilling/boring machine",
        "Drilling|Boring",
        2000,
        76,
        20000,
        None,
        "https://news.thomasnet.com/companystory/world-leaders-in-the-field-of-subsea-development-solutions-have-purchased-a-tbt-bw300-table-boring-machine-from-tbt-476348",
        "Table type; gundrill/BTA drilling Ø5-76 mm x 2000 mm depth; rotary table handles up to 20 t.",
    ),
    # ── Indexable Milling (machining centers) ──
    (
        "MAZAK VCN-460C",
        "Yamazaki Mazak",
        "Vertical machining center",
        "Indexable Milling",
        900,
        460,
        1000,
        None,
        "https://www.mazak.com/jp-en/products/vcn/",
        "Vertical machining center; table 900 x 460 mm; for keyway/flat milling.",
    ),
    (
        "MAZAK VCN-700D",
        "Yamazaki Mazak",
        "Vertical machining center",
        "Indexable Milling",
        1740,
        700,
        2000,
        None,
        "https://www.mazak.com/eu-en/products/vcn/",
        "Vertical machining center; table 1740 x 700 mm; large work envelope for milling.",
    ),
    (
        "DMG MORI DMU 50",
        "DMG MORI",
        "5-axis universal milling centre",
        "Indexable Milling",
        500,
        630,
        600,
        None,
        "https://www.dmgmori.com/en/products/milling/monoblock/5-axis/universal/dmu-50",
        "5-axis universal milling; travel X/Y/Z 500/450/400 mm, table Ø630 mm.",
    ),
    # ── Boring (horizontal boring mills) ──
    (
        "TOS WHN 13 CNC",
        "TOS Varnsdorf",
        "Horizontal boring mill",
        "Boring",
        2000,
        1800,
        12000,
        None,
        "https://www.machinio.au/listings/69867756-tos-whn-13-in-tortona-italy",
        "Spindle Ø130 mm, X travel 2000 mm, Y 2000 mm, Z 1250 mm; table 1800x1800 mm; workpiece up to 12 t.",
    ),
    (
        "TOS WHN 110",
        "TOS Varnsdorf",
        "Horizontal boring mill",
        "Boring",
        2000,
        1400,
        8000,
        None,
        "https://www.aucto.com/data/tos-varnsdorf-whn-110",
        "Spindle Ø110 mm, X travel 2000/3000 mm, table 1400x1600 mm; workpiece up to 8 t (rotary table).",
    ),
    # ── Thread Grinding (thread grinders) ──
    (
        "MATRIX 37 CNC",
        "Matrix Machine Tool",
        "CNC universal thread grinder",
        "Thread Grinding",
        508,
        355,
        100,
        None,
        "https://www.indiamart.com/proddetail/matrix-37-cnc-six-axis-cnc-thread-grinder-machine-for-sale-20197198862.html",
        "Max diameter ground 355 mm, max length between centres 508 mm, CNC 6-axis.",
    ),
    (
        "MATRIX 47 Type 1",
        "Matrix Machine Tool",
        "Heavy-duty universal thread grinder",
        "Thread Grinding",
        1066,
        355,
        200,
        None,
        "https://www.indiamart.com/proddetail/47-type-1-matrix-heavy-duty-universal-thread-grinder-22984788091.html",
        "Max diameter admitted 355 mm, max length between centres 1066 mm.",
    ),
    (
        "MATRIX THW-3080",
        "Matrix Machine Tool",
        "CNC worm thread grinding machine",
        "Thread Grinding",
        800,
        300,
        80,
        None,
        "https://www.matrix-machine.tw/en/THW2080en.html",
        "Max workpiece OD 300 mm, distance between centres 800 mm, table load 80 kg.",
    ),
    # ── Balancing (dynamic balancing machines) ──
    (
        "SCHENCK HM2",
        "Schenck RoTec",
        "Horizontal balancing machine",
        "Balancing",
        1600,
        360,
        50,
        None,
        "https://www.schenck.cn/productclass_285/productdetail_512.shtml",
        "Hard-bearing horizontal balancer; rotor up to 50 kg, journal Ø7-50 mm, bearing distance up to 1600 mm.",
    ),
    (
        "SCHENCK HM4",
        "Schenck RoTec",
        "Horizontal balancing machine",
        "Balancing",
        1750,
        1600,
        1500,
        None,
        "https://www.directindustry.com/prod/schenck-rotec-gmbh/product-14345-423556.html",
        "Hard-bearing horizontal balancer; rotor up to 1500 kg, max diameter 1600 mm, journal Ø12-200 mm.",
    ),
    (
        "SCHENCK HM40",
        "Schenck RoTec",
        "Horizontal balancing machine",
        "Balancing",
        1750,
        1600,
        3000,
        None,
        "https://www.directindustry.com/prod/schenck-rotec-gmbh/product-14345-423556.html",
        "Hard-bearing horizontal balancer; rotor up to 3000 kg, max diameter 1600 mm, journal Ø15-240 mm.",
    ),
    # ── Shot Blasting (shot blast machines) ──
    (
        "WHEELABRATOR WS-4/210",
        "Wheelabrator",
        "Spinner-hanger shot blast machine",
        "Shot Blasting",
        1000,
        635,
        900,
        None,
        "https://betadiecasting.com/machine/used-wheelbrator-shot-blaster-sinner-hanger-ws-4-210-5008/",
        "Spinner hanger; work opening 1650x1040 mm, max work height 1000 mm, max load 900 kg per hook.",
    ),
    (
        "WHEELABRATOR CT-2-45/7-430",
        "Wheelabrator",
        "Continuous through-feed shot blast machine",
        "Shot Blasting",
        700,
        700,
        250,
        None,
        "https://www.wheelabratorgroup.com/wheelblast-equipment/rotary-table-tas-rt/",
        "Continuous apron conveyor; part size diagonal up to 700 mm, for all-side descaling/peening.",
    ),
]

# Common capability added to existing lathe rows
LATHE_PROCESSES = "Threading|Taper Turning|Grooving"


def _max_dia(rod, chuck):
    vals = [v for v in (rod, chuck) if isinstance(v, (int, float))]
    return max(vals) if vals else None


def main():
    wb = load_workbook(XLSX)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]

    def row_index(name):
        # openpyxl cell(row, col) uses a 1-based column number
        return headers.index(name) + 1

    # ── 1. Add Threading/Taper/Grooving capability to existing lathes ──
    lathe_updated = 0
    for r in range(2, ws.max_row + 1):
        machine_type = ws.cell(r, row_index("Machine type")).value
        if not (machine_type and "Lathe" in str(machine_type)):
            continue
        tlen = ws.cell(r, row_index("Turning length")).value
        rod = ws.cell(r, row_index("Max. turning diameter rod.")).value
        chuck = ws.cell(r, row_index("Max. turning diameter chuck.")).value
        if not isinstance(tlen, (int, float)):
            continue
        # Do not overwrite rows that already have Supported processes
        if ws.cell(r, row_index("Supported processes")).value:
            continue
        ws.cell(r, row_index("Supported processes")).value = LATHE_PROCESSES
        ws.cell(r, row_index("Max workpiece length")).value = float(tlen)
        ws.cell(r, row_index("Max workpiece length (Unit)")).value = "mm"
        dia = _max_dia(rod, chuck)
        if dia:
            ws.cell(r, row_index("Max workpiece diameter")).value = float(dia)
            ws.cell(r, row_index("Max workpiece diameter (Unit)")).value = "mm"
        ws.cell(
            r, row_index("Capability notes")
        ).value = "CNC lathe; supports threading, taper turning and grooving."
        lathe_updated += 1

    # ── 2. Append the new real machine rows ──
    added = 0
    existing = {ws.cell(r, row_index("Designation")).value for r in range(2, ws.max_row + 1)}
    for designation, mfr, mtype, procs, length, dia, weight, module, url, note in NEW_ROWS:
        if designation in existing:
            print(f"  [skip] {designation} already exists")
            continue
        new_row = {
            "Designation": designation,
            "Unique identifier": designation,
            "Manufacturer": mfr,
            "Owner": "SystemProductCosting",
            "Capital Asset Classification": "MC",
            "Technology Classification": "Machining",
            "Machine type": mtype,
            "Supported processes": procs,
            "Max workpiece length": float(length),
            "Max workpiece length (Unit)": "mm",
            "Max workpiece diameter": float(dia),
            "Max workpiece diameter (Unit)": "mm",
            "Max workpiece weight": float(weight) if weight else None,
            "Max workpiece weight (Unit)": "kg",
            "Max gear module": float(module) if module else None,
            "Max gear module (Unit)": "mm",
            "Capability source URL": url,
            "Capability notes": note,
            "Price input mode": "Linked",
        }
        ws.append([new_row.get(h) for h in headers])
        added += 1

    wb.save(XLSX)
    print(f"Updated lathes with threading/taper/grooving: {lathe_updated}")
    print(f"Added new real machines: {added}")
    print(f"Total rows now: {ws.max_row - 1}")


if __name__ == "__main__":
    main()
